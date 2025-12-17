# -*- coding: utf-8 -*-
from flask import Blueprint, request, jsonify, send_file
import sys
import os
import io
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
from models import Template, User
from routes.auth import token_required

template_bp = Blueprint('template', __name__)

@template_bp.route('/templates', methods=['GET'])
@token_required
def get_templates():
    """获取所有模板"""
    templates = Template.get_all()
    return jsonify({'success': True, 'data': templates})

@template_bp.route('/templates/<center>', methods=['GET'])
@token_required
def get_template_by_center(center):
    """获取指定中心的模板"""
    team = request.args.get('team')
    template = Template.find_by_center(center, team)
    
    if not template:
        return jsonify({'success': False, 'message': '模板不存在'}), 404
    
    return jsonify({'success': True, 'data': template})

@template_bp.route('/templates', methods=['POST'])
@token_required
def create_or_update_template():
    """创建或更新模板（管理员）"""
    user = User.find_by_id(request.user_id)
    if user['role'] != 'admin':
        return jsonify({'success': False, 'message': '权限不足'}), 403
    
    data = request.get_json()
    
    center = data.get('center')
    team = data.get('team')
    name = data.get('name', f'{center}工时模板')
    fields = data.get('fields', [])
    
    if not center:
        return jsonify({'success': False, 'message': '中心名称不能为空'}), 400
    
    Template.create_or_update(center, team, name, fields)
    
    return jsonify({'success': True, 'message': '模板保存成功'})

@template_bp.route('/my-template', methods=['GET'])
@token_required
def get_my_template():
    """获取当前用户所属中心的模板"""
    user = User.find_by_id(request.user_id)
    center = user['center']
    team = user['team']
    
    template = Template.find_by_center(center, team)
    
    if not template:
        template = Template.find_by_center(center, None)
    
    if not template:
        return jsonify({
            'success': True,
            'data': {
                'center': center,
                'team': team,
                'name': '默认模板',
                'fields': [
                    {'key': 'category', 'label': '工作类别', 'type': 'select', 'required': True, 'options': ['日常工作', '项目工作', '会议', '其他']},
                    {'key': 'hours', 'label': '工时', 'type': 'number', 'required': True, 'min': 0, 'max': 24},
                    {'key': 'description', 'label': '工作描述', 'type': 'text', 'required': False}
                ]
            }
        })
    
    return jsonify({'success': True, 'data': template})


# ==================== Excel 模版上传/下载 API ====================

def parse_frontend_templates():
    """获取前端模版配置（直接定义，避免解析 TypeScript 的复杂性）"""
    
    # ==================== 业务管理及合规检测中心 ====================
    businessManagementCategories = [
        {'value': '_1检测相关_常规', 'label': '检测相关_常规', 'children': [
            {'value': '1.1 跨境收单业务合规检测', 'label': '1.1 跨境收单业务合规检测'},
            {'value': '1.2 整体合规/法务工作机制检测', 'label': '1.2 整体合规/法务工作机制检测'},
            {'value': '1.3 2024年支付合规团队自查复验', 'label': '1.3 2024年支付合规团队自查复验'},
            {'value': '1.4 金融消保长效工作机制后续检测', 'label': '1.4 金融消保长效工作机制后续检测'},
            {'value': '1.5 梧桐稳智平台 采购与合规性检测', 'label': '1.5 梧桐稳智平台 采购与合规性检测'},
            {'value': '1.6 PA OKR/BSC review', 'label': '1.6 PA OKR/BSC review'},
            {'value': '1.7 境外机构制裁检测', 'label': '1.7 境外机构制裁检测'},
        ]},
        {'value': '_2检测相关_快速', 'label': '检测相关_快速', 'children': [
            {'value': '2.1 非金融平台治理检测', 'label': '2.1 非金融平台治理检测'},
            {'value': '2.2 涉俄制裁风险管理与应对有效性检测', 'label': '2.2 涉俄制裁风险管理与应对有效性检测'},
            {'value': '2.3 跨团队、多主体评估流程检测', 'label': '2.3 跨团队、多主体评估流程检测'},
            {'value': '2.4 金融持牌主体反洗钱检测', 'label': '2.4 金融持牌主体反洗钱检测'},
            {'value': '2.5 金融主体投诉管理工作检测', 'label': '2.5 金融主体投诉管理工作检测'},
            {'value': '2.6 圳星/Tenpay Global 反洗钱整体检测', 'label': '2.6 圳星/Tenpay Global 反洗钱整体检测'},
            {'value': '2.7 服务商商户不能同一验证', 'label': '2.7 服务商商户不能同一验证'},
        ]},
        {'value': '_3业务管理相关_业务战略总结', 'label': '业务管理相关_业务战略总结', 'children': [
            {'value': '3.1 OKR、BSC会议', 'label': '3.1 OKR、BSC会议'},
            {'value': '3.2 部门年度BSC/OKR制定、调整', 'label': '3.2 部门年度BSC/OKR制定、调整'},
            {'value': '3.3 部门战略工作汇报、总结', 'label': '3.3 部门战略工作汇报、总结'},
            {'value': '3.4 集团、部门层面各类业务信息总结报送', 'label': '3.4 集团、部门层面各类业务信息总结报送'},
        ]},
        {'value': '_4业务管理相关_项目跟进', 'label': '业务管理相关_项目跟进', 'children': [
            {'value': '4.1 正向价值项目', 'label': '4.1 正向价值项目'},
            {'value': '4.2 反洗钱制裁相关项目', 'label': '4.2 反洗钱制裁相关项目'},
            {'value': '4.3 支付相关项目', 'label': '4.3 支付相关项目'},
            {'value': '4.4 金融理财相关项目', 'label': '4.4 金融理财相关项目'},
            {'value': '4.5 消保相关项目', 'label': '4.5 消保相关项目'},
            {'value': '4.6 境外主体合规管理', 'label': '4.6 境外主体合规管理'},
            {'value': '4.7 金融科技相关', 'label': '4.7 金融科技相关'},
            {'value': '4.8 检测项目后续跟踪', 'label': '4.8 检测项目后续跟踪'},
        ]},
        {'value': '_5公共_流程机制', 'label': '公共_流程机制', 'children': [
            {'value': '5.1 跨部门/团队流程梳理', 'label': '5.1 跨部门/团队流程梳理'},
            {'value': '5.2 VOC量化评估', 'label': '5.2 VOC量化评估'},
            {'value': '5.3 内部工作机制优化', 'label': '5.3 内部工作机制优化'},
        ]},
        {'value': '_6公共_部门公共事务支持', 'label': '公共_部门公共事务支持', 'children': [
            {'value': '6.1 各部门管理例会及业务会议', 'label': '6.1 各部门管理例会及业务会议'},
            {'value': '6.2 预算管理', 'label': '6.2 预算管理'},
            {'value': '6.3 IT管理', 'label': '6.3 IT管理'},
            {'value': '6.4 管理类总结', 'label': '6.4 管理类总结'},
            {'value': '6.5 其他', 'label': '6.5 其他'},
        ]},
        {'value': '_7公共_执业管理', 'label': '公共_执业管理', 'children': [
            {'value': '7.1 知识分享-公司内', 'label': '7.1 知识分享-公司内'},
            {'value': '7.2 知识分享-公司外', 'label': '7.2 知识分享-公司外'},
            {'value': '7.3 参加内、外部培训', 'label': '7.3 参加内、外部培训'},
            {'value': '7.4 金融合规培训体系升级', 'label': '7.4 金融合规培训体系升级'},
            {'value': '7.5 内部知识管理体系搭建', 'label': '7.5 内部知识管理体系搭建'},
            {'value': '7.6 AI信息赋能能力建设', 'label': '7.6 AI信息赋能能力建设'},
            {'value': '7.7 内部员工成长体系建设', 'label': '7.7 内部员工成长体系建设'},
            {'value': '7.8 团队影响力建设', 'label': '7.8 团队影响力建设'},
        ]},
        {'value': '_8管理_仅leader使用', 'label': '管理_仅leader使用', 'children': [
            {'value': '8.1 员工培养及辅导', 'label': '8.1 员工培养及辅导'},
            {'value': '8.2 跨团队管理事务沟通', 'label': '8.2 跨团队管理事务沟通'},
            {'value': '8.3 管理类会议组织及参会', 'label': '8.3 管理类会议组织及参会'},
            {'value': '8.4 组织氛围建设', 'label': '8.4 组织氛围建设'},
            {'value': '8.5 团队招聘', 'label': '8.5 团队招聘'},
        ]},
        {'value': '_9其他', 'label': '其他', 'children': [
            {'value': '9.1 工时填写', 'label': '9.1 工时填写'},
            {'value': '9.2 团队/部门例会', 'label': '9.2 团队/部门例会'},
            {'value': '9.3 团队日报/周报/月报填写', 'label': '9.3 团队日报/周报/月报填写'},
            {'value': '9.4 差旅路途时间', 'label': '9.4 差旅路途时间'},
            {'value': '9.5 商务招待（非工作时间段）', 'label': '9.5 商务招待（非工作时间段）'},
        ]},
    ]
    
    # 标签（BSC/OKR/Others）及其关键任务
    businessManagementTags = [
        {'value': '_BSC', 'label': 'BSC', 'children': [
            {'value': '检测机制持续优化', 'label': '检测机制持续优化'},
            {'value': '正向价值机制维护与运行', 'label': '正向价值机制维护与运行'},
            {'value': '金融科技板块跟进与运行', 'label': '金融科技板块跟进与运行'},
            {'value': '五部门战略工作机制运营维护', 'label': '五部门战略工作机制运营维护'},
            {'value': '产品清单全盘梳理', 'label': '产品清单全盘梳理'},
            {'value': '校招生轮岗机制维护', 'label': '校招生轮岗机制维护'},
            {'value': '知识管理优化', 'label': '知识管理优化'},
            {'value': '预算管理机制维护', 'label': '预算管理机制维护'},
            {'value': '团队专业能力培养及影响力提升', 'label': '团队专业能力培养及影响力提升'},
            {'value': '风险合规预警平台搭建', 'label': '风险合规预警平台搭建'},
            {'value': '研发项目与IT系统搭建', 'label': '研发项目与IT系统搭建'},
            {'value': '金融合规培训活动运营', 'label': '金融合规培训活动运营'},
            {'value': 'AI信息赋能能力持续建设', 'label': 'AI信息赋能能力持续建设'},
            {'value': '金融职能支持部门信息上报运营', 'label': '金融职能支持部门信息上报运营'},
            {'value': '金融职能支持部门跨团队流程制定及优化', 'label': '金融职能支持部门跨团队流程制定及优化'},
            {'value': '金融职能支持部门日常运营支持', 'label': '金融职能支持部门日常运营支持'},
            {'value': '反诈合规履职工作框架及机制明确', 'label': '反诈合规履职工作框架及机制明确'},
            {'value': '消保合规框架及一号位梳理支持', 'label': '消保合规框架及一号位梳理支持'},
            {'value': '财付通一号位梳理及合规闭环流程制定', 'label': '财付通一号位梳理及合规闭环流程制定'},
            {'value': '全面支持香港钱包合规管理及监管沟通等工作', 'label': '全面支持香港钱包合规管理及监管沟通等工作'},
        ]},
        {'value': '_OKR', 'label': 'OKR', 'children': [
            {'value': '合规检测项目开展', 'label': '合规检测项目开展'},
            {'value': 'VOC量化评估体系', 'label': 'VOC量化评估体系'},
        ]},
        {'value': '_Others', 'label': 'Others', 'children': []},
    ]
    
    businessManagementWorkTypes = [
        {'value': '项目调研、访谈、资料查阅学习等工作', 'label': '项目调研、访谈、资料查阅学习等工作'},
        {'value': '项目方案讨论、制定', 'label': '项目方案讨论、制定'},
        {'value': '项目执行相关的数据调取/分析、抽样工作', 'label': '项目执行相关的数据调取/分析、抽样工作'},
        {'value': '项目执行结果分析、总结、汇报工作', 'label': '项目执行结果分析、总结、汇报工作'},
        {'value': '项目跟踪', 'label': '项目跟踪'},
        {'value': '部门各类会议支持', 'label': '部门各类会议支持（包括会议前期准备、会议召开、会议总结等工作）'},
        {'value': '部门各类公共支持事务答疑', 'label': '部门各类公共支持事务答疑'},
        {'value': '部门拉通类项目推进', 'label': '部门拉通类项目推进'},
        {'value': '部门内/跨部门知识分享', 'label': '部门内/跨部门知识分享'},
        {'value': '团队、部门目标管理工作', 'label': '团队、部门目标管理工作'},
        {'value': '参与工作相关的各类培训', 'label': '参与工作相关的各类培训'},
        {'value': '其他', 'label': '其他'},
    ]
    
    # ==================== 投资法务中心 ====================
    investmentLegalGroups = [
        {'value': '1组', 'label': '1组'},
        {'value': '2组', 'label': '2组'},
        {'value': '3组', 'label': '3组'},
        {'value': '4组', 'label': '4组'},
        {'value': '5组', 'label': '5组'},
        {'value': '6组', 'label': '6组'},
    ]
    
    investmentLegalCategories = [
        {'value': 'Investment Related - M&A Deal', 'label': 'Investment Related - M&A Deal', 'children': [
            {'value': 'Drafting/reviewing/revising legal documents', 'label': 'Drafting/reviewing/revising legal documents'},
            {'value': 'Internal/external discussion on/negotiation of legal documents', 'label': 'Internal/external discussion on/negotiation of legal documents'},
            {'value': 'Conducting LDD/reviewing LDD results', 'label': 'Conducting LDD/reviewing LDD results'},
            {'value': 'Conducting legal researches', 'label': 'Conducting legal researches'},
            {'value': 'Interacting with authorities/regulators', 'label': 'Interacting with authorities/regulators'},
            {'value': 'Reverting to inquiries', 'label': 'Reverting to inquiries'},
            {'value': 'Others', 'label': 'Others'},
        ]},
        {'value': 'Investment Related - IPO', 'label': 'Investment Related - IPO', 'children': [
            {'value': 'Drafting/reviewing/revising legal documents', 'label': 'Drafting/reviewing/revising legal documents'},
            {'value': 'Internal/external discussion on/negotiation of legal documents', 'label': 'Internal/external discussion on/negotiation of legal documents'},
            {'value': 'Conducting LDD/reviewing LDD results', 'label': 'Conducting LDD/reviewing LDD results'},
            {'value': 'Conducting legal researches', 'label': 'Conducting legal researches'},
            {'value': 'Interacting with authorities/regulators', 'label': 'Interacting with authorities/regulators'},
            {'value': 'Reverting to inquiries', 'label': 'Reverting to inquiries'},
            {'value': 'Others', 'label': 'Others'},
        ]},
        {'value': 'Investment Related - Corporate Matter', 'label': 'Investment Related - Corporate Matter', 'children': [
            {'value': 'Drafting/reviewing/revising legal documents', 'label': 'Drafting/reviewing/revising legal documents'},
            {'value': 'Internal/external discussion on/negotiation of legal documents', 'label': 'Internal/external discussion on/negotiation of legal documents'},
            {'value': 'Conducting LDD/reviewing LDD results', 'label': 'Conducting LDD/reviewing LDD results'},
            {'value': 'Conducting legal researches', 'label': 'Conducting legal researches'},
            {'value': 'Interacting with authorities/regulators', 'label': 'Interacting with authorities/regulators'},
            {'value': 'Reverting to inquiries', 'label': 'Reverting to inquiries'},
            {'value': 'Others', 'label': 'Others'},
        ]},
        {'value': 'Non-Investment Related - Other Departments', 'label': 'Non-Investment Related - Other Departments', 'children': [
            {'value': 'CTD（合规交易部 - 公司事务及国际监管法务中心）', 'label': 'CTD（合规交易部 - 公司事务及国际监管法务中心）'},
            {'value': 'AMLSC（反洗钱与制裁合规部）', 'label': 'AMLSC（反洗钱与制裁合规部）'},
            {'value': 'FLCD（金融法律合规部）', 'label': 'FLCD（金融法律合规部）'},
            {'value': 'FPAD（金融业务公共事务部）', 'label': 'FPAD（金融业务公共事务部）'},
            {'value': 'FCRPD（集团金融消费者权益保护部）', 'label': 'FCRPD（集团金融消费者权益保护部）'},
            {'value': 'Others', 'label': 'Others'},
        ]},
        {'value': 'Public - Infrastructure & Guidance', 'label': 'Public - Infrastructure & Guidance', 'children': [
            {'value': 'Preparing policies/procedures/guidance/booklet alert', 'label': 'Preparing policies/procedures/guidance/booklet alert'},
            {'value': 'Establishing/improving IT systems', 'label': 'Establishing/improving IT systems'},
            {'value': 'Updating internal templates', 'label': 'Updating internal templates'},
            {'value': 'Participating training sessions', 'label': 'Participating training sessions'},
            {'value': 'Others', 'label': 'Others'},
        ]},
        {'value': 'Public - Knowledge Accumulation & Sharing', 'label': 'Public - Knowledge Accumulation & Sharing', 'children': [
            {'value': 'Preparing knowhow/memo/client alert', 'label': 'Preparing knowhow/memo/client alert'},
            {'value': 'Providing training sessions/knowledge sharing', 'label': 'Providing training sessions/knowledge sharing'},
            {'value': 'Conducting CTD Documents Collation Project (iWiki)', 'label': 'Conducting CTD Documents Collation Project (iWiki)'},
            {'value': "Relevant matters related to engagement of law firms/team's legal fees", 'label': "Relevant matters related to engagement of law firms/team's legal fees"},
            {'value': 'Others', 'label': 'Others'},
        ]},
        {'value': 'Public - Others', 'label': 'Public - Others', 'children': [
            {'value': 'Participating regular team meetings', 'label': 'Participating regular team meetings'},
            {'value': 'Recording/reviewing deals updates/weekly time reports/other team summaries', 'label': 'Recording/reviewing deals updates/weekly time reports/other team summaries'},
            {'value': 'Supporting GCTSM-related matters', 'label': 'Supporting GCTSM-related matters'},
            {'value': 'Interacting with regulators', 'label': 'Interacting with regulators'},
            {'value': 'Others', 'label': 'Others'},
        ]},
    ]
    
    investmentLegalBSCTags = [
        {'value': 'BSC', 'label': 'BSC'},
        {'value': 'Others', 'label': 'Others'},
    ]
    
    investmentLegalBSCItems = [
        {'value': '境内投融资法务服务', 'label': '境内投融资法务服务 Domestic investment legal services'},
        {'value': '境外投融资法务服务', 'label': '境外投融资法务服务 Overseas investment legal services'},
        {'value': '国际监管趋势监测、预判', 'label': '国际监管趋势监测、预判 Monitoring and forecasting international regulatory trends'},
        {'value': '境内投融资资本市场监管发展与重大政策争取与风险应对', 'label': '境内投融资资本市场监管发展与重大政策争取与风险应对'},
        {'value': '执业管理平台体系完善', 'label': '执业管理平台体系完善 Perfection of the professional management platform'},
        {'value': '优化Hotdesk和专家组制度', 'label': '优化Hotdesk和专家组制度 Optimization of Hotdesk and expert panel system'},
        {'value': '精细化管理工具搭建', 'label': '精细化管理工具搭建 Adoption of refined management tools'},
    ]
    
    # ==================== 公司及国际金融事务中心 ====================
    internationalFinanceVirtualGroups = [
        {'value': 'Listing Rules and Corporate Governance', 'label': 'Listing Rules and Corporate Governance'},
        {'value': 'Group Financing', 'label': 'Group Financing'},
        {'value': 'International Financial', 'label': 'International Financial'},
        {'value': 'Non-CTD (兼岗)', 'label': 'Non-CTD (兼岗)'},
    ]
    
    internationalFinanceInternalClients = [
        {'value': 'MA', 'label': 'MA'},
        {'value': 'Comsec', 'label': 'Comsec'},
        {'value': 'HR', 'label': 'HR'},
        {'value': 'Tax', 'label': 'Tax'},
        {'value': 'Admin/Operations', 'label': 'Admin/Operations'},
        {'value': 'GCTSM', 'label': 'GCTSM'},
        {'value': 'Treasury', 'label': 'Treasury'},
        {'value': 'CSIG', 'label': 'CSIG'},
        {'value': 'IEGG', 'label': 'IEGG'},
        {'value': 'WXG', 'label': 'WXG'},
        {'value': 'WPHK', 'label': 'WPHK'},
        {'value': 'WPMY', 'label': 'WPMY'},
        {'value': 'Finance', 'label': 'Finance'},
        {'value': 'FiT (ZX/TG)', 'label': 'FiT (ZX/TG)'},
        {'value': 'Midas', 'label': 'Midas'},
        {'value': 'Tai Fung', 'label': 'Tai Fung'},
        {'value': 'Fusion Bank', 'label': 'Fusion Bank'},
        {'value': 'FuSure', 'label': 'FuSure'},
        {'value': 'Group matter', 'label': 'Group matter'},
        {'value': 'Others', 'label': 'Others'},
        {'value': 'N/A', 'label': 'N/A'},
    ]
    
    internationalFinanceCategories = [
        {'value': '_BSC', 'label': 'BSC', 'children': [
            {'value': 'CTD-0101 境外金融业务资质申请', 'label': 'CTD-0101 境外金融业务资质申请 (国际金融业务牌照申请)'},
            {'value': 'CTD-0105 集团融资支持', 'label': 'CTD-0105 集团融资支持 Group financing'},
            {'value': 'CTD-0106 国际监管趋势监测、预判 (研究国际监管环境)', 'label': 'CTD-0106 国际监管趋势监测、预判 (研究国际监管环境和趋势)'},
            {'value': 'CTD-0106 国际监管趋势监测、预判 (海外高风险法律)', 'label': 'CTD-0106 国际监管趋势监测、预判 (海外高风险法律的合规工作)'},
            {'value': 'CTD-0201 境外主体/办公室合规管理', 'label': 'CTD-0201 境外主体/办公室合规管理'},
            {'value': 'CTD-0202 国际人力资源合规管理', 'label': 'CTD-0202 国际人力资源合规管理 Overseas HR support'},
            {'value': 'CTD-0203 国际业务合规管理', 'label': 'CTD-0203 国际业务合规管理 Overseas businesses/products compliance'},
            {'value': 'CTD-0204 集团上市合规风险管理', 'label': 'CTD-0204 集团上市合规风险管理 Listing Rules compliance'},
            {'value': 'CTD-0205 反洗钱和制裁合规支持(境外)', 'label': 'CTD-0205 反洗钱和制裁合规支持(境外) AML and Sanctions compliance'},
            {'value': 'CTD-0209 境外金融持牌主体业务合规管理', 'label': 'CTD-0209 境外金融持牌主体业务合规管理'},
            {'value': 'CTD-0210 海外投资合规管理', 'label': 'CTD-0210 海外投资合规管理 Overseas investment compliance (CFIUS)'},
            {'value': 'CTD-0301 国际业务监管应对', 'label': 'CTD-0301 国际业务监管应对 International regulatory requests/queries/responses'},
            {'value': 'CTD-0401 执业管理平台体系完善', 'label': 'CTD-0401 执业管理平台体系完善 (智能化工作开发及运用)'},
            {'value': 'CTD-0402 优化Hotdesk制度', 'label': 'CTD-0402 优化Hotdesk制度 Optimising Hotdesks'},
            {'value': 'CTD-0403 精细化管理工具搭建', 'label': 'CTD-0403 精细化管理工具搭建 (优化完善VOC工作记录工具)'},
        ]},
        {'value': '_Others', 'label': 'Others', 'children': [
            {'value': '团队内部会议 Team meetings', 'label': '团队内部会议 Team meetings'},
            {'value': '账单处理 Billing', 'label': '账单处理 Billing'},
            {'value': '绩效统计（工时和项目数据）', 'label': '绩效统计（工时和项目数据）Timesheet and matters recording'},
        ]},
        {'value': '_Management', 'label': 'Management', 'children': [
            {'value': '管理例会 Management meetings', 'label': '管理例会 Management meetings'},
            {'value': '预算规划 Budget planning', 'label': '预算规划 Budget planning'},
            {'value': '人才管理 Human resources', 'label': '人才管理 Human resources'},
            {'value': '工作总结/汇报', 'label': '工作总结/汇报 Internal reportings / achievement summary'},
            {'value': '一般管理 General management', 'label': '一般管理 General management'},
        ]},
        {'value': '_Non_CTD_兼岗', 'label': 'Non-CTD (兼岗)', 'children': [
            {'value': 'Global Research (GPA)', 'label': 'Global Research (GPA)'},
            {'value': 'International Regulatory Policy (GPA)', 'label': 'International Regulatory Policy (GPA)'},
        ]},
    ]
    
    internationalFinanceWorkCategories = [
        {'value': 'Drafting/reviewing/commenting on documents', 'label': 'Drafting/reviewing/commenting on documents'},
        {'value': 'Discussing with internal legal team/internal stakeholders/external counsels', 'label': 'Discussing with internal legal team/internal stakeholders/external counsels'},
        {'value': 'Negotiating/discussing with external parties', 'label': 'Negotiating/discussing with external parties'},
        {'value': 'Conducting legal analysis and research', 'label': 'Conducting legal analysis and research'},
        {'value': 'Preparing internal know-how / team admin work', 'label': 'Preparing internal know-how / team admin work'},
        {'value': 'Preparing internal memo / policy / procedures / guidance etc.', 'label': 'Preparing internal memo / policy / procedures / guidance etc.'},
        {'value': 'Participating in training sessions', 'label': 'Participating in training sessions'},
        {'value': 'Internal/cross-teams knowledge sharing', 'label': 'Internal/cross-teams knowledge sharing'},
        {'value': 'Participating training sessions/team meetings', 'label': 'Participating training sessions/team meetings'},
        {'value': 'Liaising with regulators or preparing written response to regulators', 'label': 'Liaising with regulators or preparing written response to regulators'},
        {'value': 'Others', 'label': 'Others'},
    ]
    
    # 构建 keyTask 选项（从 tag 的 children 提取，带上 parentValue 标记 BSC/OKR）
    keyTask_options = []
    for tag_opt in businessManagementTags:
        if tag_opt.get('children'):
            for child in tag_opt['children']:
                keyTask_options.append({
                    'value': child['value'],
                    'label': child['label'],
                    'parentValue': tag_opt['value']  # _BSC 或 _OKR
                })
    
    # 构建 task 选项（从 category 的 children 提取，带上 parentValue）
    task_options = []
    for cat_opt in businessManagementCategories:
        if cat_opt.get('children'):
            for child in cat_opt['children']:
                task_options.append({
                    'value': child['value'],
                    'label': child['label'],
                    'parentValue': cat_opt['value']
                })
    
    # 投资法务中心：workCategory 选项（从 category 的 children 提取）
    investmentWorkCategory_options = []
    for cat_opt in investmentLegalCategories:
        if cat_opt.get('children'):
            for child in cat_opt['children']:
                investmentWorkCategory_options.append({
                    'value': child['value'],
                    'label': child['label'],
                    'parentValue': cat_opt['value']
                })
    
    # 公司及国际金融事务中心：item 选项（从 tag 的 children 提取）
    internationalItem_options = []
    for tag_opt in internationalFinanceCategories:
        if tag_opt.get('children'):
            for child in tag_opt['children']:
                internationalItem_options.append({
                    'value': child['value'],
                    'label': child['label'],
                    'parentValue': tag_opt['value']
                })
    
    # 构建模版
    templates = [
        {
            'center': '业务管理及合规检测中心',
            'team': None,
            'fields': [
                {'key': 'category', 'label': '事项分类', 'type': 'select', 'required': True, 'placeholder': '请选择事项分类', 'options': businessManagementCategories},
                {'key': 'task', 'label': '工作任务', 'type': 'select', 'required': True, 'placeholder': '请选择工作任务', 'options': task_options, 'parentField': 'category'},
                {'key': 'tag', 'label': '标签', 'type': 'select', 'required': True, 'placeholder': '请选择标签', 'options': businessManagementTags},
                {'key': 'keyTask', 'label': '关键任务', 'type': 'select', 'required': False, 'placeholder': '请选择关键任务', 'options': keyTask_options, 'parentField': 'tag', 'conditionalRequired': {'dependsOn': 'tag', 'when': ['_BSC', '_OKR']}},
                {'key': 'hours', 'label': '小时数', 'type': 'number', 'required': True, 'placeholder': '请输入小时数', 'min': 0, 'max': 24},
                {'key': 'workType', 'label': '工作类型', 'type': 'select', 'required': True, 'placeholder': '请选择工作类型', 'options': businessManagementWorkTypes},
                {'key': 'description', 'label': '描述', 'type': 'text', 'required': False, 'placeholder': '简要描述工作内容（可选）'},
            ]
        },
        {
            'center': '投资法务中心',
            'team': None,
            'fields': [
                {'key': 'sourcePath', 'label': '小组 (Team)', 'type': 'select', 'required': True, 'placeholder': '请选择组别', 'options': investmentLegalGroups},
                {'key': 'category', 'label': 'Deal/Matter Category (事务类别)', 'type': 'select', 'required': True, 'placeholder': '请选择事务类别', 'options': investmentLegalCategories},
                {'key': 'dealName', 'label': 'Deal/Matter Name (项目名称)', 'type': 'text', 'required': False, 'placeholder': '请输入TIM系统中的公司名称', 'conditionalRequired': {'dependsOn': 'category', 'when': ['Investment Related - M&A Deal', 'Investment Related - IPO', 'Investment Related - Corporate Matter']}},
                {'key': 'bscTag', 'label': 'BSC Tag (是否BSC事项)', 'type': 'select', 'required': True, 'placeholder': '请确认是否为CTD-ILC BSC事项', 'options': investmentLegalBSCTags},
                {'key': 'bscItem', 'label': 'BSC Item (BSC事项)', 'type': 'select', 'required': False, 'placeholder': '如为BSC事项，请选择具体项目', 'options': investmentLegalBSCItems, 'conditionalRequired': {'dependsOn': 'bscTag', 'when': 'BSC'}},
                {'key': 'hours', 'label': 'Hours (小时数)', 'type': 'number', 'required': True, 'placeholder': '请输入小时数', 'min': 0, 'max': 24},
                {'key': 'workCategory', 'label': 'Work Category (工作类别)', 'type': 'select', 'required': True, 'placeholder': '请选择工作类别', 'options': investmentWorkCategory_options, 'parentField': 'category'},
                {'key': 'description', 'label': 'Narrative (工作描述)', 'type': 'text', 'required': False, 'placeholder': '简要描述工作内容（可选）'},
            ]
        },
        {
            'center': '公司及国际金融事务中心',
            'team': None,
            'fields': [
                {'key': 'virtualGroup', 'label': 'Virtual Group', 'type': 'select', 'required': True, 'placeholder': '请选择Virtual Group', 'options': internationalFinanceVirtualGroups},
                {'key': 'internalClient', 'label': 'Internal Client', 'type': 'select', 'required': True, 'placeholder': '请选择Internal Client（如无需求方请选N/A）', 'options': internationalFinanceInternalClients},
                {'key': 'tag', 'label': 'Tag', 'type': 'select', 'required': True, 'placeholder': '请确认该事务是否为BSC事项', 'options': internationalFinanceCategories},
                {'key': 'item', 'label': 'Item', 'type': 'select', 'required': True, 'placeholder': '请选择对应事项', 'options': internationalItem_options, 'parentField': 'tag'},
                {'key': 'hours', 'label': 'Hours', 'type': 'number', 'required': True, 'placeholder': '请输入小时数', 'min': 0, 'max': 24},
                {'key': 'workCategory', 'label': 'Work Category', 'type': 'select', 'required': True, 'placeholder': '请选择工作类别', 'options': internationalFinanceWorkCategories},
                {'key': 'description', 'label': 'Narrative (Optional)', 'type': 'text', 'required': False, 'placeholder': '简要描述工作内容'},
            ]
        },
    ]
    
    return templates


@template_bp.route('/templates/download', methods=['GET'])
@token_required
def download_template_excel():
    """下载模版 Excel 文件"""
    user = User.find_by_id(request.user_id)
    if user['role'] != 'admin':
        return jsonify({'success': False, 'message': '权限不足'}), 403
    
    templates = Template.get_all()
    
    # 如果数据库中没有模版，从前端配置文件读取
    if not templates:
        templates = parse_frontend_templates()
    
    wb = Workbook()
    
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    
    ws_fields = wb.active
    ws_fields.title = '字段定义'
    
    field_headers = ['中心名称', '团队名称', '字段Key', '字段名称', '字段类型', '是否必填', '父字段Key', '条件必填依赖字段', '条件必填触发值', 'placeholder', '最小值', '最大值']
    for col, header in enumerate(field_headers, 1):
        cell = ws_fields.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border
    
    row = 2
    for template in templates:
        center = template.get('center', '')
        team = template.get('team', '')
        for field in template.get('fields', []):
            ws_fields.cell(row=row, column=1, value=center).border = thin_border
            ws_fields.cell(row=row, column=2, value=team or '').border = thin_border
            ws_fields.cell(row=row, column=3, value=field.get('key', '')).border = thin_border
            ws_fields.cell(row=row, column=4, value=field.get('label', '')).border = thin_border
            ws_fields.cell(row=row, column=5, value=field.get('type', 'text')).border = thin_border
            ws_fields.cell(row=row, column=6, value='是' if field.get('required') else '否').border = thin_border
            ws_fields.cell(row=row, column=7, value=field.get('parentField', '')).border = thin_border
            cond_req = field.get('conditionalRequired', {})
            ws_fields.cell(row=row, column=8, value=cond_req.get('dependsOn', '')).border = thin_border
            when_val = cond_req.get('when', '')
            if isinstance(when_val, list):
                when_val = '|'.join(when_val)
            ws_fields.cell(row=row, column=9, value=when_val).border = thin_border
            ws_fields.cell(row=row, column=10, value=field.get('placeholder', '')).border = thin_border
            ws_fields.cell(row=row, column=11, value=field.get('min', '')).border = thin_border
            ws_fields.cell(row=row, column=12, value=field.get('max', '')).border = thin_border
            row += 1
    
    for col in range(1, 13):
        ws_fields.column_dimensions[chr(64 + col)].width = 18
    
    ws_options = wb.create_sheet('选项配置')
    option_headers = ['中心名称', '字段Key', '父选项值', '选项值', '选项显示名']
    for col, header in enumerate(option_headers, 1):
        cell = ws_options.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border
    
    row = 2
    for template in templates:
        center = template.get('center', '')
        for field in template.get('fields', []):
            field_key = field.get('key', '')
            options = field.get('options', [])
            has_parent_field = field.get('parentField')  # 是否是子字段（如 task 依赖 category）
            
            def write_options(opts, parent_value=''):
                nonlocal row
                for opt in opts:
                    if isinstance(opt, dict):
                        # 优先使用选项自带的 parentValue（用于 keyTask 等从父字段 children 提取的选项）
                        actual_parent = opt.get('parentValue', parent_value)
                        
                        # 如果该字段有 parentField，说明它是子字段，只写叶子节点（带 parentValue 的选项）
                        # 如果没有 parentField，且选项有 children，则只写父级选项（不写 children）
                        if has_parent_field:
                            # 子字段：写入所有选项（它们应该已经是扁平化的，带 parentValue）
                            ws_options.cell(row=row, column=1, value=center).border = thin_border
                            ws_options.cell(row=row, column=2, value=field_key).border = thin_border
                            ws_options.cell(row=row, column=3, value=actual_parent).border = thin_border
                            ws_options.cell(row=row, column=4, value=opt.get('value', '')).border = thin_border
                            ws_options.cell(row=row, column=5, value=opt.get('label', '')).border = thin_border
                            row += 1
                        else:
                            # 父字段：只写父级选项，不递归写 children
                            ws_options.cell(row=row, column=1, value=center).border = thin_border
                            ws_options.cell(row=row, column=2, value=field_key).border = thin_border
                            ws_options.cell(row=row, column=3, value=actual_parent).border = thin_border
                            ws_options.cell(row=row, column=4, value=opt.get('value', '')).border = thin_border
                            ws_options.cell(row=row, column=5, value=opt.get('label', '')).border = thin_border
                            row += 1
                            # 不递归 children，因为 children 应该属于子字段
                    else:
                        ws_options.cell(row=row, column=1, value=center).border = thin_border
                        ws_options.cell(row=row, column=2, value=field_key).border = thin_border
                        ws_options.cell(row=row, column=3, value=parent_value).border = thin_border
                        ws_options.cell(row=row, column=4, value=opt).border = thin_border
                        ws_options.cell(row=row, column=5, value=opt).border = thin_border
                        row += 1
            
            if options:
                write_options(options)
    
    for col in range(1, 6):
        ws_options.column_dimensions[chr(64 + col)].width = 25
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True, download_name='template_config.xlsx')


@template_bp.route('/templates/upload', methods=['POST'])
@token_required
def upload_template_excel():
    """上传 Excel 文件更新模版配置"""
    user = User.find_by_id(request.user_id)
    if user['role'] != 'admin':
        return jsonify({'success': False, 'message': '权限不足'}), 403
    
    if 'file' not in request.files:
        return jsonify({'success': False, 'message': '请选择文件'}), 400
    
    file = request.files['file']
    if not file.filename.endswith(('.xlsx', '.xls')):
        return jsonify({'success': False, 'message': '请上传 Excel 文件 (.xlsx 或 .xls)'}), 400
    
    try:
        wb = load_workbook(file)
        
        if '字段定义' not in wb.sheetnames:
            return jsonify({'success': False, 'message': 'Excel 文件缺少"字段定义"工作表'}), 400
        
        ws_fields = wb['字段定义']
        ws_options = wb['选项配置'] if '选项配置' in wb.sheetnames else None
        
        templates_data = {}
        
        for row in ws_fields.iter_rows(min_row=2, values_only=True):
            if not row[0]:
                continue
            
            center = str(row[0]).strip()
            team = str(row[1]).strip() if row[1] else None
            field_key = str(row[2]).strip() if row[2] else ''
            field_label = str(row[3]).strip() if row[3] else ''
            field_type = str(row[4]).strip() if row[4] else 'text'
            required = str(row[5]).strip() == '是' if row[5] else False
            parent_field = str(row[6]).strip() if row[6] else None
            cond_depends = str(row[7]).strip() if row[7] else None
            cond_when = str(row[8]).strip() if row[8] else None
            placeholder = str(row[9]).strip() if len(row) > 9 and row[9] else None
            min_val = row[10] if len(row) > 10 and row[10] else None
            max_val = row[11] if len(row) > 11 and row[11] else None
            
            if not field_key:
                continue
            
            key = (center, team)
            if key not in templates_data:
                templates_data[key] = {'fields': [], 'name': f'{center}工时模板'}
            
            field = {'key': field_key, 'label': field_label, 'type': field_type, 'required': required}
            
            if parent_field:
                field['parentField'] = parent_field
            if placeholder:
                field['placeholder'] = placeholder
            if min_val is not None:
                field['min'] = float(min_val) if min_val else 0
            if max_val is not None:
                field['max'] = float(max_val) if max_val else 24
            
            if cond_depends and cond_when:
                when_values = cond_when.split('|') if '|' in cond_when else cond_when
                field['conditionalRequired'] = {'dependsOn': cond_depends, 'when': when_values}
            
            templates_data[key]['fields'].append(field)
        
        if ws_options:
            options_data = {}
            for row in ws_options.iter_rows(min_row=2, values_only=True):
                if not row[0] or not row[1]:
                    continue
                center = str(row[0]).strip()
                field_key = str(row[1]).strip()
                parent_value = str(row[2]).strip() if row[2] else ''
                opt_value = str(row[3]).strip() if row[3] else ''
                opt_label = str(row[4]).strip() if len(row) > 4 and row[4] else opt_value
                
                if not opt_value:
                    continue
                
                key = (center, field_key)
                if key not in options_data:
                    options_data[key] = {}
                if parent_value not in options_data[key]:
                    options_data[key][parent_value] = []
                options_data[key][parent_value].append({'value': opt_value, 'label': opt_label})
            
            for (center, team), template_info in templates_data.items():
                for field in template_info['fields']:
                    opt_key = (center, field['key'])
                    if opt_key in options_data:
                        opts = options_data[opt_key]
                        def build_options(parent=''):
                            result = []
                            if parent in opts:
                                for opt in opts[parent]:
                                    option = {'value': opt['value'], 'label': opt['label']}
                                    children = build_options(opt['value'])
                                    if children:
                                        option['children'] = children
                                    result.append(option)
                            return result
                        field['options'] = build_options('')
        
        saved_count = 0
        for (center, team), template_info in templates_data.items():
            Template.create_or_update(center, team, template_info['name'], template_info['fields'])
            saved_count += 1
        
        return jsonify({'success': True, 'message': f'成功更新 {saved_count} 个模版', 'data': {'count': saved_count}})
        
    except Exception as e:
        return jsonify({'success': False, 'message': f'解析文件失败: {str(e)}'}), 400


@template_bp.route('/templates/sample', methods=['GET'])
@token_required
def download_sample_template():
    """下载空白模版示例 Excel"""
    user = User.find_by_id(request.user_id)
    if user['role'] != 'admin':
        return jsonify({'success': False, 'message': '权限不足'}), 403
    
    wb = Workbook()
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    example_fill = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    
    ws_fields = wb.active
    ws_fields.title = '字段定义'
    
    field_headers = ['中心名称', '团队名称', '字段Key', '字段名称', '字段类型', '是否必填', '父字段Key', '条件必填依赖字段', '条件必填触发值', 'placeholder', '最小值', '最大值']
    for col, header in enumerate(field_headers, 1):
        cell = ws_fields.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border
    
    example_fields = [
        ['投资法务中心', '', 'category', 'Deal/Matter Category', 'select', '是', '', '', '', '请选择事务类别', '', ''],
        ['投资法务中心', '', 'dealName', 'Deal/Matter Name', 'text', '否', '', 'category', 'Investment Related - M&A Deal|Investment Related - IPO', '请输入项目名称', '', ''],
        ['投资法务中心', '', 'hours', 'Hours', 'number', '是', '', '', '', '请输入小时数', '0', '24'],
        ['投资法务中心', '', 'workCategory', 'Work Category', 'cascader', '是', 'category', '', '', '请选择工作类别', '', ''],
    ]
    
    for row_idx, row_data in enumerate(example_fields, 2):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws_fields.cell(row=row_idx, column=col_idx, value=value)
            cell.fill = example_fill
            cell.border = thin_border
    
    for col in range(1, 13):
        ws_fields.column_dimensions[chr(64 + col)].width = 20
    
    ws_options = wb.create_sheet('选项配置')
    option_headers = ['中心名称', '字段Key', '父选项值', '选项值', '选项显示名']
    for col, header in enumerate(option_headers, 1):
        cell = ws_options.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border
    
    example_options = [
        ['投资法务中心', 'category', '', 'Investment Related - M&A Deal', 'Investment Related - M&A Deal'],
        ['投资法务中心', 'category', '', 'Investment Related - IPO', 'Investment Related - IPO'],
        ['投资法务中心', 'category', '', 'Non-Investment Related', 'Non-Investment Related'],
        ['投资法务中心', 'workCategory', 'Investment Related - M&A Deal', 'Due Diligence', 'Due Diligence'],
        ['投资法务中心', 'workCategory', 'Investment Related - M&A Deal', 'Negotiation', 'Negotiation'],
    ]
    
    for row_idx, row_data in enumerate(example_options, 2):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws_options.cell(row=row_idx, column=col_idx, value=value)
            cell.fill = example_fill
            cell.border = thin_border
    
    for col in range(1, 6):
        ws_options.column_dimensions[chr(64 + col)].width = 30
    
    ws_help = wb.create_sheet('填写说明')
    help_content = [
        '模版配置说明', '',
        '一、字段定义 Sheet',
        '- 中心名称：必填，如"投资法务中心"、"业务管理及合规检测中心"',
        '- 团队名称：可选，如果某团队有特殊模版可填写',
        '- 字段Key：必填，英文标识，如 category、hours、dealName',
        '- 字段名称：必填，显示给用户的名称',
        '- 字段类型：select(下拉选择)、cascader(级联选择)、text(文本)、number(数字)',
        '- 是否必填：是/否',
        '- 父字段Key：级联选择时填写，表示依赖哪个字段',
        '- 条件必填依赖字段：当某字段值满足条件时才必填',
        '- 条件必填触发值：多个值用 | 分隔', '',
        '二、选项配置 Sheet',
        '- 中心名称：与字段定义中的中心名称对应',
        '- 字段Key：与字段定义中的字段Key对应',
        '- 父选项值：用于级联选择，表示该选项属于哪个父选项',
        '- 选项值：选项的实际值',
        '- 选项显示名：显示给用户的文字',
    ]
    
    for row_idx, content in enumerate(help_content, 1):
        cell = ws_help.cell(row=row_idx, column=1, value=content)
        if row_idx == 1:
            cell.font = Font(bold=True, size=14)
        elif content.startswith('一、') or content.startswith('二、'):
            cell.font = Font(bold=True)
    
    ws_help.column_dimensions['A'].width = 80
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True, download_name='template_sample.xlsx')
